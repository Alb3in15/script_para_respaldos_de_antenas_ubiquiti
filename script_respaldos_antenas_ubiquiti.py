import paramiko
import openpyxl
from os import mkdir
from datetime import datetime

def respaldar_antena(ip, usuario, contrasenia,directorio):
    try:
        cliente_ssh = paramiko.SSHClient()
        cliente_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        cliente_ssh.connect(ip, username=usuario, password=contrasenia, timeout=5)
        
        # Comandos para respaldar la configuración de la antena Ubiquiti
        comando_respaldo = "cat /tmp/system.cfg"
        stdin, stdout, stderr = cliente_ssh.exec_command(comando_respaldo)
        configuracion = stdout.read().decode()

        # Crear un archivo de respaldo con la configuración
        fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_archivo = f"respaldo_{ip}_{fecha_actual}.cfg"
        mi_directorio = directorio
        with open(f"{mi_directorio}/{nombre_archivo}", "w") as archivo:
            archivo.write(configuracion)

        print(f"Respaldo de la antena en {ip} creado con éxito: {nombre_archivo}")
        # Puedes enviar comandos SSH utilizando cliente_ssh.exec_command() o utilizar alguna otra biblioteca específica para interactuar con la antena Ubiquiti
        
        print(f"Respaldo generado para la antena con IP {ip}")
        
    except paramiko.AuthenticationException:
        print(f"Error de autenticación para la antena con IP {ip}. Verifica las credenciales proporcionadas.")
    except paramiko.SSHException as e:
        print(f"Error SSH para la antena con IP {ip}: {e}")
    except paramiko.ssh_exception.NoValidConnectionsError:
        print(f"No se pudo establecer conexión SSH con la antena con IP {ip}")
    except Exception as e:
        print(f"Error desconocido para la antena con IP {ip}: {e}")
    finally:
        cliente_ssh.close()  # Asegúrate de cerrar la conexión SSH

# Lee el archivo Excel con los rangos de IP
def leer_archivo_excel(nombre_archivo):
    wb = openpyxl.load_workbook(nombre_archivo)
    hoja = wb.active

    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    directorio = mkdir("Aps_Ubiquiti_Guabo_y_Rivera_" + fecha_actual)
    # Supongamos que las IPs están en la columna A (desde la fila 2 hasta la última fila)
    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=1, values_only=True):
        ip = fila[0]
        usuario_ssh = "usuario"
        contrasenia_ssh = "contrasenia"
        # Realiza el respaldo de la antena Ubiquiti
        respaldar_antena(ip, usuario_ssh, contrasenia_ssh, f"Aps_Ubiquiti_Guabo_y_Rivera_{fecha_actual}")

# Ejemplo de uso
nombre_archivo_excel = "APs.xlsx"  # Reemplaza con el nombre de tu archivo Excel
leer_archivo_excel(nombre_archivo_excel)
